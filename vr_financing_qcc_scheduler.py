"""
基于企查查融资信息核查 API 的 VR/XR 公司融资抓取脚本。

特点：
- 只使用企查查开放平台的「融资信息核查」接口（ApiCode 950），不做全网网页抓取；
- 按多组与 VR/XR 相关的关键词查询融资事件；
- 将结果整理成统一格式写入 Excel 文件，文件名精确到分钟，存放在 `vr_reports_qcc/` 目录；
- 可作为“结构化主数据源”，与 `vr_financing_scheduler.py` 的全网搜索版本互补。
"""
import datetime
import time
from pathlib import Path
from typing import List, Dict

import openpyxl
from openpyxl.utils import get_column_letter
import schedule

from config import QCC_API_KEY, QCC_SECRET_KEY
from data_sources.qcc_client import fetch_vr_financing_qcc


def fetch_vr_financing_data_qcc() -> List[Dict[str, str]]:
    """
    使用企查查 API 拉取 VR/XR 相关公司融资记录。

    NOTE：
    - 需要在 .env 中配置 QCC_API_KEY / QCC_SECRET_KEY；
    - 实际返回的数据取决于企查查是否有对应记录。
    """
    if not QCC_API_KEY or not QCC_SECRET_KEY:
        print("未配置 QCC_API_KEY 或 QCC_SECRET_KEY，无法使用企查查融资接口。")
        return []

    print("使用企查查 API 拉取 VR/XR 相关融资数据...")
    try:
        records = fetch_vr_financing_qcc(QCC_API_KEY, QCC_SECRET_KEY)
    except Exception as e:
        print(f"企查查 API 调用失败：{e}")
        return []

    if not records:
        print("企查查未返回任何 VR/XR 相关融资记录。")
    else:
        print(f"企查查共返回 {len(records)} 条 VR/XR 融资记录。")
    return records


def export_to_excel(records: List[Dict[str, str]], filename: Path) -> None:
    """将融资数据写入 Excel 文件（覆盖写）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VR融资情况(QCC)"

    headers = ["公司名称", "轮次", "融资金额", "融资时间", "行业", "来源"]
    ws.append(headers)

    if not records:
        print("本次企查查未返回数据，将生成只包含表头的空 Excel 模板。")
    else:
        for item in records:
            row = [item.get(h, "") for h in headers]
            ws.append(row)

    for idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for cell in ws[get_column_letter(idx)]:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

    filename.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(filename))
    print(f"已写入 Excel 文件：{filename}")


def job_once() -> None:
    """执行一次：企查查 + Excel 输出，若已配置飞书则上传到云文档。"""
    print(f"[{datetime.datetime.now().isoformat(sep=' ', timespec='seconds')}] 开始从企查查拉取 VR/XR 融资数据...")
    data = fetch_vr_financing_data_qcc()
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    reports_dir = Path(__file__).resolve().parent / "vr_reports_qcc"
    filename = reports_dir / f"vr_financing_qcc_{ts}.xlsx"
    export_to_excel(data, filename)
    try:
        from data_sources.feishu_upload import upload_if_configured
        upload_if_configured(filename)
    except Exception as e:
        print(f"飞书上传跳过或失败：{e}")
    print(f"[{datetime.datetime.now().isoformat(sep=' ', timespec='seconds')}] 本次企查查任务完成。")


def run_scheduler() -> None:
    """
    启动定时任务（企查查版）：
    - 启动时立即执行一次；
    - 之后每 60 分钟执行一次（企查查接口有调用成本，频率不宜太高，可按需调整）。
    """
    schedule.clear()
    schedule.every(60).minutes.do(job_once)
    print("企查查定时任务已启动：每 60 分钟拉取一次 VR/XR 融资数据并生成 Excel（vr_reports_qcc）。按 Ctrl+C 可退出。")

    # 先立即执行一次，方便测试
    job_once()

    while True:
        next_run = schedule.next_run()
        if next_run is None:
            sleep_seconds = 30
        else:
            now = datetime.datetime.now()
            delta = (next_run - now).total_seconds()
            sleep_seconds = max(5, int(delta))

        for remaining in range(sleep_seconds, 0, -1):
            print(f"\r[QCC] 距离下次拉取还有 {remaining:4d} 秒", end="", flush=True)
            time.sleep(1)
        print()

        schedule.run_pending()


if __name__ == "__main__":
    run_scheduler()

