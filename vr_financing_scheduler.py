"""
定时搜索国内 VR / A 轮融资情况，并整理成 Excel。

当前版本（更系统化）：
- 不是依赖单一数据源，而是「全网搜索 + 站点筛选 + AI 抽取」组合：
  - 多组关键词在搜索引擎上全网搜索；
  - 对结果按域名做白名单 / 黑名单过滤（优先新闻/行业媒体等站点）；
  - 抓取网页正文，喂给 LLM，让其抽取结构化融资事件；
  - 聚合去重后生成 Excel。
- 脚本一运行就立刻抓取一次并生成一份 Excel，之后每分钟再跑一次，并打印倒计时。
- 每一份 Excel 的文件名精确到分钟，不会覆盖原文件，统一保存在 `vr_reports/` 文件夹。
"""
import datetime
import json
import time
from pathlib import Path
from typing import List, Dict, Any, Set

import openpyxl
from openpyxl.utils import get_column_letter
import schedule
import requests
from bs4 import BeautifulSoup
from openai import OpenAI

from config import OPENAI_API_KEY, OPENAI_BASE_URL, OPENAI_MODEL


# 域名白名单：优先保留这些站点（行业/科技/新闻等）
ALLOWED_DOMAINS = [
    "36kr.com",
    "36kr.net",
    "huxiu.com",
    "tech.163.com",
    "finance.sina.com.cn",
    "www.sina.com.cn",
    "www.jiemian.com",
    "www.yicai.com",
    "www.thepaper.cn",
    "www.cs.com.cn",
]

# 域名黑名单：尽量过滤掉不适合做结构化融资抽取的站点
BLOCKED_DOMAINS = [
    "zhihu.com",
    "zhidao.baidu.com",
    "baike.baidu.com",
    "baidu.com/s",
]

# 运行时动态累积的黑名单：多次访问失败或总是无有效信息的站点会加入这里
RUNTIME_BLOCKED_DOMAINS: Set[str] = set()
DOMAIN_NO_INFO_COUNTS: Dict[str, int] = {}

# 永久黑名单文件路径（当前脚本同级目录）
RUNTIME_BLACKLIST_PATH = Path(__file__).resolve().parent / "vr_runtime_blacklist.txt"


def _domain(url: str) -> str:
    try:
        host = url.split("//", 1)[-1].split("/", 1)[0]
        return host.lower()
    except Exception:
        return ""


# 启动时尝试加载历史运行生成的永久黑名单，加入到运行时黑名单中
try:
    if RUNTIME_BLACKLIST_PATH.exists():
        for line in RUNTIME_BLACKLIST_PATH.read_text(encoding="utf-8").splitlines():
            dom = line.strip()
            if dom and not dom.startswith("#"):
                RUNTIME_BLOCKED_DOMAINS.add(dom.lower())
        if RUNTIME_BLOCKED_DOMAINS:
            print(f"已从 vr_runtime_blacklist.txt 加载 {len(RUNTIME_BLOCKED_DOMAINS)} 个永久黑名单域名。")
except Exception as e:
    print(f"加载永久黑名单失败：{e}")


def _get_llm_client() -> OpenAI:
    kwargs: Dict[str, Any] = {"api_key": OPENAI_API_KEY}
    if OPENAI_BASE_URL:
        kwargs["base_url"] = OPENAI_BASE_URL
    return OpenAI(**kwargs)


def _plan_search(client: OpenAI) -> List[Dict[str, str]]:
    """
    让 LLM 先根据目标“VR/XR 相关公司融资事件”生成一组搜索计划，
    然后再按计划去搜索。

    期望返回结构：[{ "query": "...", "industry": "VR/AR/XR" }, ...]
    """
    system_prompt = (
        "你是一个搜索策略专家，目标是帮助用户通过中文互联网找到"
        "「与 VR / AR / XR / 元宇宙 相关公司的各轮次股权融资事件」的新闻或报道。"
        "你只需要设计搜索引擎用的查询语句，不需要去实际搜索。"
    )
    user_prompt = (
        "请设计一组用于必应(Bing)的中文搜索查询，用来尽可能全面地找到"
        "与 VR/AR/XR/元宇宙 相关公司的融资新闻（不限轮次，包含天使轮、A/B/C轮、战略融资等）。\n\n"
        "要求：\n"
        "1. 返回一个 JSON 数组，每个元素是一个对象，字段：query（搜索关键词）、industry（统一填 \"VR/AR/XR\"）。\n"
        "2. query 要贴近真实媒体标题写法，可以包含：\n"
        "   - “完成融资”“获融资”“新一轮融资”“融资消息”等表述；\n"
        "   - VR/AR/XR/虚拟现实/元宇宙 等行业词；\n"
        "   - 可以适当加入 2023/2024/2025 等年份，偏向近几年；\n"
        "3. 数量 8~12 条之间，不要重复或仅做极细微改动的查询；\n"
        "4. 严格只输出 JSON，不要有任何解释或额外文字。"
    )

    try:
        resp = client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.3,
        )
        content = resp.choices[0].message.content or "[]"
        data = _extract_json_from_text(content)
        if isinstance(data, dict):
            data = [data]
        if not isinstance(data, list):
            raise ValueError("LLM 返回格式不是数组")
        plan: List[Dict[str, str]] = []
        for item in data:
            if not isinstance(item, dict):
                continue
            q = str(item.get("query") or "").strip()
            if not q:
                continue
            industry = str(item.get("industry") or "VR/AR/XR").strip() or "VR/AR/XR"
            plan.append({"query": q, "industry": industry})
        # 兜底：至少要有几条
        if not plan:
            raise ValueError("计划为空")
        return plan
    except Exception as e:
        print(f"LLM 生成搜索计划失败，将回退到内置默认计划: {e}")
        # 内置一个简单的兜底计划，避免完全不可用
        fallback = [
            {"query": "VR 公司 完成 融资", "industry": "VR/AR/XR"},
            {"query": "虚拟现实 公司 获 新一轮 融资", "industry": "VR/AR/XR"},
            {"query": "XR AR VR 创业公司 融资 消息", "industry": "VR/AR/XR"},
            {"query": "元宇宙 公司 完成 融资", "industry": "VR/AR/XR"},
        ]
        return fallback


def _search_bing(query: str, limit: int = 10) -> List[Dict[str, str]]:
    url = "https://www.bing.com/search"
    # 尝试一次多拿一些结果，并偏向中文
    params = {
        "q": query,
        "count": 30,          # 请求更多结果（具体生效与否取决于 Bing）
        "setLang": "zh-cn",
        "cc": "CN",
    }
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/123.0 Safari/537.36"
        )
    }

    try:
        resp = requests.get(url, params=params, headers=headers, timeout=10)
        resp.raise_for_status()
    except Exception as e:
        print(f"请求必应失败（{query}）: {e}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    results: List[Dict[str, str]] = []
    raw_results: List[Dict[str, str]] = []
    for li in soup.select("li.b_algo"):
        a = li.select_one("h2 a")
        if not a:
            continue
        title = a.get_text(strip=True)
        link = a.get("href", "")
        snippet_el = li.select_one(".b_caption p")
        snippet = snippet_el.get_text(strip=True) if snippet_el else ""
        raw_results.append({"title": title, "link": link, "snippet": snippet})

    # 先按白名单过滤其一批结果，若为空，再用黑名单过滤后的通用结果兜底
    allowed: List[Dict[str, str]] = []
    fallback: List[Dict[str, str]] = []
    for item in raw_results:
        dom = _domain(item["link"])
        # 静态黑名单和运行时黑名单都跳过
        if any(bad in dom for bad in BLOCKED_DOMAINS) or dom in RUNTIME_BLOCKED_DOMAINS:
            continue
        if any(allowed_dom in dom for allowed_dom in ALLOWED_DOMAINS):
            allowed.append(item)
        else:
            fallback.append(item)

    ordered = allowed + fallback
    for item in ordered:
        results.append(item)
        if len(results) >= limit:
            break
    return results


def _fetch_page_text(url: str) -> str:
    try:
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/123.0 Safari/537.36"
            )
        }
        resp = requests.get(url, headers=headers, timeout=10)
        resp.raise_for_status()
    except Exception as e:
        print(f"获取网页失败：{url} - {e}")
        dom = _domain(url)
        if dom:
            RUNTIME_BLOCKED_DOMAINS.add(dom)
            print(f"已将站点加入运行时黑名单：{dom}")
        return ""

    soup = BeautifulSoup(resp.text, "html.parser")
    text = soup.get_text(separator="\n", strip=True)
    # 避免内容过长导致 token 爆炸，这里截断前 8000 字符
    return text[:8000]


def _extract_json_from_text(text: str) -> Any:
    """尽量从模型返回中提取 JSON（容错处理，引号、Markdown 包裹等）。"""
    text = text.strip()
    # 去掉 ```json ``` 包裹
    if text.startswith("```"):
        text = text.strip("`")
        if text.lower().startswith("json"):
            text = text[4:].strip()
    # 找到第一个 '[' 或 '{'
    start_idx = len(text)
    for ch in ("[", "{"):
        i = text.find(ch)
        if i != -1:
            start_idx = min(start_idx, i)
    if start_idx == len(text):
        raise ValueError("未找到 JSON 起始符号")
    text = text[start_idx:]
    # 从尾部开始找最后一个 ']' 或 '}'
    end_idx = -1
    for ch in ("]", "}"):
        j = text.rfind(ch)
        if j > end_idx:
            end_idx = j
    if end_idx == -1:
        raise ValueError("未找到 JSON 结束符号")
    text = text[: end_idx + 1]
    return json.loads(text)


def _llm_structured_extract(client: OpenAI, page_text: str, url: str, industry: str) -> List[Dict[str, str]]:
    if not page_text:
        return []
    system_prompt = (
        "你是一个专业的一级市场投融资分析助手。"
        "现在给你一篇网页的正文内容，请你从中提取**所有与国内 VR / AR / XR 等相关公司的各轮次股权融资事件**。"
        "输出严格的 JSON 数组，每个元素是一个对象，字段包括：\n"
        "- 公司名称（string）\n"
        "- 轮次（string，如 天使轮、A轮、A+轮、B轮、C轮、战略融资 等）\n"
        "- 融资金额（string，例如 “数千万元人民币”、“1亿元人民币”等）\n"
        "- 融资时间（string，尽量用 YYYY-MM 或 YYYY-MM-DD；若无法确定，可写“未知”）\n"
        "- 行业（string，例如 VR、AR、XR，若无法确定，可写“未知”）\n"
        "- 来源（string，写简短来源说明，可包含该网页 URL）\n"
        "如果网页中没有任何相关融资事件，请输出空数组 []。"
    )
    user_prompt = (
        f"网页 URL：{url}\n\n"
        "网页正文内容如下（可能不完整）：\n"
        "--------------------\n"
        f"{page_text}\n"
        "--------------------\n\n"
        "请按照系统提示，提取相关融资事件，并只输出 JSON。"
    )
    try:
        resp = client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.1,
        )
    except Exception as e:
        print(f"调用 LLM 解析网页失败：{url} - {e}")
        return []

    content = resp.choices[0].message.content or ""
    try:
        data = _extract_json_from_text(content)
    except Exception as e:
        print(f"解析 LLM 返回 JSON 失败：{e}，原始内容截断：{content[:200]}...")
        return []

    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        return []

    records: List[Dict[str, str]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        records.append(
            {
                "公司名称": str(item.get("公司名称") or item.get("company") or "未知"),
                "轮次": str(item.get("轮次") or item.get("round") or "未知"),
                "融资金额": str(item.get("融资金额") or item.get("amount") or "未知"),
                "融资时间": str(item.get("融资时间") or item.get("date") or "未知"),
                "行业": str(item.get("行业") or industry or "未知"),
                "来源": str(item.get("来源") or url),
            }
        )
    return records


def fetch_vr_financing_data() -> List[Dict[str, str]]:
    """
    全网搜索 + 站点筛选 + AI 抽取：
    1. 用多组 VR + A 轮关键词在搜索引擎上检索；
    2. 按域名白名单/黑名单筛选站点（优先新闻/行业媒体）；
    3. 抓取网页正文，喂给 LLM 做结构化融资事件抽取；
    4. 聚合去重后返回统一格式列表。
    """
    client = _get_llm_client()
    all_records: Dict[str, Dict[str, str]] = {}
    search_report: List[Dict[str, str]] = []

    # 先让 LLM 生成一份搜索计划
    search_plan = _plan_search(client)
    print("本轮搜索计划：")
    for cfg in search_plan:
        print(f"  - query={cfg['query']} | industry={cfg['industry']}")
    print()

    for cfg in search_plan:
        query = cfg["query"]
        industry = cfg["industry"]
        print(f"正在搜索：{query}")
        results = _search_bing(query, limit=20)
        if not results:
            search_report.append(
                {
                    "query": query,
                    "url": "",
                    "domain": "",
                    "status": "搜索无结果",
                }
            )
            continue

        for r in results:
            url = r["link"]
            dom = _domain(url)
            print(f"  解析网页：{url}")
            page_text = _fetch_page_text(url)
            if not page_text:
                search_report.append(
                    {
                        "query": query,
                        "url": url,
                        "domain": dom,
                        "status": "获取失败或内容为空",
                    }
                )
                continue
            extracted = _llm_structured_extract(client, page_text, url, industry)
            if not extracted:
                # 统计“有正文但无有效融资信息”的站点，超过阈值则加入运行时黑名单
                if dom:
                    DOMAIN_NO_INFO_COUNTS[dom] = DOMAIN_NO_INFO_COUNTS.get(dom, 0) + 1
                    if DOMAIN_NO_INFO_COUNTS[dom] >= 3 and dom not in RUNTIME_BLOCKED_DOMAINS:
                        RUNTIME_BLOCKED_DOMAINS.add(dom)
                        print(f"站点多次无有效融资信息，已加入运行时黑名单：{dom}")
                search_report.append(
                    {
                        "query": query,
                        "url": url,
                        "domain": dom,
                        "status": "已抓取正文，未提取到融资事件",
                    }
                )
                continue
            search_report.append(
                {
                    "query": query,
                    "url": url,
                    "domain": dom,
                    "status": f"提取到 {len(extracted)} 条融资事件",
                }
            )
            for item in extracted:
                key = f"{item.get('公司名称','')}|{item.get('轮次','')}|{item.get('融资时间','')}"
                if key in all_records:
                    continue
                all_records[key] = item

    records = list(all_records.values())
    if not records:
        print("本次未能从网页中提取到任何 VR/XR 融资记录。")
    else:
        print(f"本次共从网页中提取出 {len(records)} 条 VR/XR 融资记录。")

    # 打印搜索报告，帮助你理解这一轮到底搜到了什么
    print("\n===== 本轮搜索报告 =====")
    if not search_report:
        print("本次搜索未能从搜索引擎拿到任何可用网页结果。")
    else:
        for entry in search_report:
            q = entry["query"]
            dom = entry["domain"] or "-"
            status = entry["status"]
            url = entry["url"] or "-"
            print(f"[{q}] {dom} | {status} | {url}")
    print("===== 搜索报告结束 =====\n")

    # 在一轮搜索结束后，把当前运行时黑名单持久化到文件中，形成“永久黑名单”
    try:
        if RUNTIME_BLOCKED_DOMAINS:
            text = "\n".join(sorted(RUNTIME_BLOCKED_DOMAINS))
            RUNTIME_BLACKLIST_PATH.write_text(text, encoding="utf-8")
            print(f"已更新永久黑名单文件：{RUNTIME_BLACKLIST_PATH}，共 {len(RUNTIME_BLOCKED_DOMAINS)} 个域名。")
    except Exception as e:
        print(f"写入永久黑名单失败：{e}")

    return records


def export_to_excel(records: List[Dict[str, str]], filename: Path) -> None:
    """将融资数据写入 Excel 文件（覆盖写）。

    即使本次没有抓到任何记录，也会生成带表头的空模板，方便你查看结构。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VR融资情况"

    # 表头顺序
    headers = ["公司名称", "轮次", "融资金额", "融资时间", "行业", "来源"]
    ws.append(headers)

    if not records:
        # 没有数据时，仅写入表头，方便你确认列定义
        print("本次没有抓到任何记录，将生成只包含表头的空 Excel 模板。")
    else:
        for item in records:
            row = [item.get(h, "") for h in headers]
            ws.append(row)

    # 自动列宽（简单根据最大长度估算）
    for idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for cell in ws[get_column_letter(idx)]:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

    # 确保父目录存在
    filename.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(filename))
    print(f"已写入 Excel 文件：{filename}")


def job_once() -> None:
    """执行一次：抓取数据并写入 Excel，若已配置飞书则上传到云文档。"""
    print(f"[{datetime.datetime.now().isoformat(sep=' ', timespec='seconds')}] 开始抓取 VR / A 轮融资数据...")
    data = fetch_vr_financing_data()
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    reports_dir = Path(__file__).resolve().parent / "vr_reports"
    filename = reports_dir / f"vr_financing_{ts}.xlsx"
    export_to_excel(data, filename)
    try:
        from data_sources.feishu_upload import upload_if_configured
        upload_if_configured(filename)
    except Exception as e:
        print(f"飞书上传跳过或失败：{e}")
    print(f"[{datetime.datetime.now().isoformat(sep=' ', timespec='seconds')}] 本次任务完成。")


def run_scheduler() -> None:
    """
    启动定时任务：
    - 启动时立刻执行一次 job_once（方便测试）。
    - 之后每 1 分钟执行一次 job_once，并在控制台打印倒计时。
    - 真正上线时可以改回每天某个固定时间。
    """
    # 启动时先同步本地已导出的 Excel 到飞书（若已配置飞书）
    try:
        from data_sources.feishu_upload import sync_local_excel_to_feishu
        sync_local_excel_to_feishu()
    except Exception as e:
        print(f"启动时飞书同步跳过或失败：{e}")

    schedule.clear()
    schedule.every(1).minutes.do(job_once)
    print("定时任务已启动：每 1 分钟抓取一次必应搜索结果并生成 VR 融资 Excel。按 Ctrl+C 可退出。")

    # 先立即执行一次，方便测试
    job_once()

    while True:
        # 计算距离下一次执行的剩余秒数
        # schedule.next_run 是一个函数，需要调用得到下次运行时间
        next_run = schedule.next_run()
        if next_run is None:
            # 理论上不会出现，没有任务时简单等 1 秒
            sleep_seconds = 1
        else:
            now = datetime.datetime.now()
            delta = (next_run - now).total_seconds()
            sleep_seconds = max(1, int(delta))  # 至少等待 1 秒

        for remaining in range(sleep_seconds, 0, -1):
            print(f"\r距离下次抓取还有 {remaining:3d} 秒", end="", flush=True)
            time.sleep(1)
        print()  # 换行

        schedule.run_pending()


if __name__ == "__main__":
    # 直接运行本文件时，默认启动定时任务
    run_scheduler()

